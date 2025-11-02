#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Blue Pharma Trading PLC - Enhanced Excel Analytics Report Generator
Provides comprehensive weekly analytics with professional formatting and charts.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, timedelta
import tempfile
import os
import logging

logger = logging.getLogger(__name__)

class ExcelAnalyticsGenerator:
    """Enhanced Excel report generator with professional formatting and charts."""
    
    def __init__(self):
        self.header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        self.data_font = Font(name='Calibri', size=11)
        self.title_font = Font(name='Calibri', size=14, bold=True)
        self.header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.wrap_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    def generate_weekly_analytics_report(self, db_manager, filename_prefix="Blue_Pharma_Weekly_Analytics"):
        """Generate comprehensive weekly analytics Excel report with professional formatting."""
        try:
            # Get weekly analytics data
            weekly_data = db_manager.get_weekly_analytics_data(8)
            
            if not weekly_data:
                logger.warning("No weekly analytics data available")
                return None
            
            # Create temporary Excel file
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
                temp_file_path = temp_file.name
            
            # Create workbook and worksheets
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create multiple sheets
            ws_weekly = wb.create_sheet("Weekly Analytics")
            ws_dashboard = wb.create_sheet("Dashboard")
            ws_summary = wb.create_sheet("Summary")
            
            # Generate weekly analytics sheet
            self._create_weekly_analytics_sheet(ws_weekly, weekly_data)
            
            # Generate dashboard sheet with charts
            self._create_dashboard_sheet(ws_dashboard, weekly_data)
            
            # Generate summary sheet
            self._create_summary_sheet(ws_summary, weekly_data)
            
            # Save the workbook
            wb.save(temp_file_path)
            wb.close()
            
            # Create final filename with timestamp
            current_date = datetime.now().strftime('%Y-%m-%d_%H-%M')
            final_filename = f"{filename_prefix}_{current_date}.xlsx"
            
            return temp_file_path, final_filename
            
        except Exception as e:
            logger.error(f"Error generating weekly analytics report: {e}", exc_info=True)
            return None, None
    
    def _create_weekly_analytics_sheet(self, ws, weekly_data):
        """Create the main weekly analytics sheet with enhanced formatting."""
        try:
            # Set sheet title
            ws.title = "Weekly Analytics"
            
            # Add report header
            ws['A1'] = "Blue Pharma Trading PLC - Weekly Analytics Report"
            ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='366092')
            ws.merge_cells('A1:J1')
            ws['A1'].alignment = self.center_alignment
            
            # Add generation date
            ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(name='Calibri', size=11, italic=True)
            ws.merge_cells('A2:J2')
            ws['A2'].alignment = self.center_alignment
            
            # Add some spacing
            ws.row_dimensions[3].height = 10
            
            # Prepare data for DataFrame
            df_data = []
            for week in weekly_data:
                df_data.append({
                    'Week Number': week.get('week_number', ''),
                    'Week Start': week.get('week_start', ''),
                    'Week End': week.get('week_end', ''),
                    'New Users': week.get('new_users', 0),
                    'Active Users': week.get('active_users', 0),
                    'Total Messages': week.get('total_messages', 0),
                    'Orders/Requests': week.get('orders_requests', 0),
                    'Revenue (ETB)': week.get('revenue', 0.0),
                    'Top User': week.get('top_user', 'N/A'),
                    'Notes': week.get('notes', '')
                })
            
            # Create DataFrame
            df = pd.DataFrame(df_data)
            
            # Add column headers starting at row 4
            headers = list(df.columns)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_idx, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.border
            
            # Add data rows
            for row_idx, (_, row) in enumerate(df.iterrows(), 5):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = self.data_font
                    cell.border = self.border
                    
                    # Special formatting for specific columns
                    if col_idx == 8:  # Revenue column
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    elif col_idx in [4, 5, 6, 7]:  # Numeric columns
                        cell.number_format = '#,##0'
                        cell.alignment = self.center_alignment
                    elif col_idx == 10:  # Notes column
                        cell.alignment = self.wrap_alignment
                    else:
                        cell.alignment = self.center_alignment
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
            # Freeze header row
            ws.freeze_panes = 'A5'
            
            # Add table formatting
            table_range = f"A4:{chr(64 + len(headers))}{4 + len(df)}"
            table = Table(displayName="WeeklyAnalyticsTable", ref=table_range)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9", 
                showFirstColumn=False,
                showLastColumn=False, 
                showRowStripes=True, 
                showColumnStripes=False
            )
            ws.add_table(table)
            
        except Exception as e:
            logger.error(f"Error creating weekly analytics sheet: {e}", exc_info=True)
    
    def _create_dashboard_sheet(self, ws, weekly_data):
        """Create dashboard sheet with charts and KPIs."""
        try:
            ws.title = "Dashboard"
            
            # Add dashboard title
            ws['A1'] = "Weekly Analytics Dashboard"
            ws['A1'].font = Font(name='Calibri', size=18, bold=True, color='366092')
            ws.merge_cells('A1:H1')
            ws['A1'].alignment = self.center_alignment
            
            # Create KPI section
            self._create_kpi_section(ws, weekly_data)
            
            # Create charts
            self._create_revenue_trend_chart(ws, weekly_data)
            self._create_user_activity_chart(ws, weekly_data)
            self._create_performance_pie_chart(ws, weekly_data)
            
        except Exception as e:
            logger.error(f"Error creating dashboard sheet: {e}", exc_info=True)
    
    def _create_kpi_section(self, ws, weekly_data):
        """Create KPI summary section on dashboard."""
        try:
            if not weekly_data:
                return
            
            # Calculate KPIs
            total_revenue = sum(week.get('revenue', 0) for week in weekly_data)
            total_new_users = sum(week.get('new_users', 0) for week in weekly_data)
            total_messages = sum(week.get('total_messages', 0) for week in weekly_data)
            total_orders = sum(week.get('orders_requests', 0) for week in weekly_data)
            avg_weekly_revenue = total_revenue / len(weekly_data) if weekly_data else 0
            
            # Latest week data
            latest_week = weekly_data[0] if weekly_data else {}
            
            # KPI headers
            kpi_row = 3
            ws[f'A{kpi_row}'] = "Key Performance Indicators (KPIs)"
            ws[f'A{kpi_row}'].font = Font(name='Calibri', size=14, bold=True)
            ws.merge_cells(f'A{kpi_row}:H{kpi_row}')
            
            # KPI values
            kpi_data = [
                ("Total Revenue (8 weeks)", f"{total_revenue:.2f} ETB"),
                ("Average Weekly Revenue", f"{avg_weekly_revenue:.2f} ETB"),
                ("Total New Users", f"{total_new_users:,}"),
                ("Total Messages", f"{total_messages:,}"),
                ("Total Orders/Requests", f"{total_orders:,}"),
                ("Latest Week Revenue", f"{latest_week.get('revenue', 0):.2f} ETB"),
                ("Latest Week Active Users", f"{latest_week.get('active_users', 0):,}"),
                ("Latest Week Messages", f"{latest_week.get('total_messages', 0):,}")
            ]
            
            # Add KPI data in a 2-column layout
            for i, (kpi_name, kpi_value) in enumerate(kpi_data):
                row = kpi_row + 2 + (i // 2)
                col_offset = 0 if i % 2 == 0 else 4
                
                # KPI name
                name_cell = ws.cell(row=row, column=1 + col_offset, value=kpi_name)
                name_cell.font = Font(name='Calibri', size=11, bold=True)
                name_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                name_cell.border = self.border
                
                # KPI value
                value_cell = ws.cell(row=row, column=2 + col_offset, value=kpi_value)
                value_cell.font = Font(name='Calibri', size=11, bold=True, color='366092')
                value_cell.border = self.border
                value_cell.alignment = Alignment(horizontal='right', vertical='center')
                
                # Merge cells for better appearance
                if col_offset == 0:
                    ws.merge_cells(f"{chr(65 + col_offset)}{row}:{chr(66 + col_offset)}{row}")
                    ws.merge_cells(f"{chr(67 + col_offset)}{row}:{chr(68 + col_offset)}{row}")
                else:
                    ws.merge_cells(f"{chr(69 + col_offset)}{row}:{chr(70 + col_offset)}{row}")
                    ws.merge_cells(f"{chr(71 + col_offset)}{row}:{chr(72 + col_offset)}{row}")
            
        except Exception as e:
            logger.error(f"Error creating KPI section: {e}", exc_info=True)
    
    def _create_revenue_trend_chart(self, ws, weekly_data):
        """Create revenue trend line chart."""
        try:
            if not weekly_data or len(weekly_data) < 2:
                return
            
            # Add chart data starting at row 15
            chart_data_row = 15
            
            # Headers
            ws[f'A{chart_data_row}'] = "Week"
            ws[f'B{chart_data_row}'] = "Revenue (ETB)"
            ws[f'C{chart_data_row}'] = "Orders"
            
            # Format headers
            for col in ['A', 'B', 'C']:
                cell = ws[f'{col}{chart_data_row}']
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.border
            
            # Add data (reverse order for chronological display)
            for i, week in enumerate(reversed(weekly_data), 1):
                data_row = chart_data_row + i
                ws[f'A{data_row}'] = week.get('week_number', f'Week {i}')
                ws[f'B{data_row}'] = week.get('revenue', 0)
                ws[f'C{data_row}'] = week.get('orders_requests', 0)
                
                # Format data cells
                for col in ['A', 'B', 'C']:
                    cell = ws[f'{col}{data_row}']
                    cell.font = self.data_font
                    cell.border = self.border
                    if col == 'B':
                        cell.number_format = '#,##0.00'
                    elif col == 'C':
                        cell.number_format = '#,##0'
                    cell.alignment = self.center_alignment
            
            # Create line chart for revenue trend
            chart = LineChart()
            chart.title = "Revenue Trend Over Time"
            chart.style = 2
            chart.y_axis.title = 'Revenue (ETB)'
            chart.x_axis.title = 'Week'
            chart.width = 15
            chart.height = 10
            
            # Add data to chart
            data_range = Reference(ws, min_col=2, min_row=chart_data_row, max_row=chart_data_row + len(weekly_data), max_col=2)
            categories = Reference(ws, min_col=1, min_row=chart_data_row + 1, max_row=chart_data_row + len(weekly_data))
            chart.add_data(data_range, titles_from_data=True)
            chart.set_categories(categories)
            
            # Position chart
            ws.add_chart(chart, "E15")
            
        except Exception as e:
            logger.error(f"Error creating revenue trend chart: {e}", exc_info=True)
    
    def _create_user_activity_chart(self, ws, weekly_data):
        """Create user activity bar chart."""
        try:
            if not weekly_data:
                return
            
            # Add chart data starting at row 30
            chart_data_row = 30
            
            # Headers
            ws[f'A{chart_data_row}'] = "Week"
            ws[f'B{chart_data_row}'] = "New Users"
            ws[f'C{chart_data_row}'] = "Active Users"
            ws[f'D{chart_data_row}'] = "Messages"
            
            # Format headers
            for col in ['A', 'B', 'C', 'D']:
                cell = ws[f'{col}{chart_data_row}']
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.border
            
            # Add data (reverse order for chronological display)
            for i, week in enumerate(reversed(weekly_data), 1):
                data_row = chart_data_row + i
                ws[f'A{data_row}'] = week.get('week_number', f'Week {i}')
                ws[f'B{data_row}'] = week.get('new_users', 0)
                ws[f'C{data_row}'] = week.get('active_users', 0)
                ws[f'D{data_row}'] = week.get('total_messages', 0)
                
                # Format data cells
                for col in ['A', 'B', 'C', 'D']:
                    cell = ws[f'{col}{data_row}']
                    cell.font = self.data_font
                    cell.border = self.border
                    if col in ['B', 'C', 'D']:
                        cell.number_format = '#,##0'
                        cell.alignment = self.center_alignment
                    else:
                        cell.alignment = self.center_alignment
            
            # Create bar chart for user activity
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "User Activity Trends"
            chart.y_axis.title = 'Count'
            chart.x_axis.title = 'Week'
            chart.width = 15
            chart.height = 10
            
            # Add data series
            new_users_data = Reference(ws, min_col=2, min_row=chart_data_row, max_row=chart_data_row + len(weekly_data))
            active_users_data = Reference(ws, min_col=3, min_row=chart_data_row, max_row=chart_data_row + len(weekly_data))
            categories = Reference(ws, min_col=1, min_row=chart_data_row + 1, max_row=chart_data_row + len(weekly_data))
            
            chart.add_data(new_users_data, titles_from_data=True)
            chart.add_data(active_users_data, titles_from_data=True)
            chart.set_categories(categories)
            
            # Position chart
            ws.add_chart(chart, "F30")
            
        except Exception as e:
            logger.error(f"Error creating user activity chart: {e}", exc_info=True)
    
    def _create_performance_pie_chart(self, ws, weekly_data):
        """Create performance breakdown pie chart."""
        try:
            if not weekly_data:
                return
            
            # Calculate performance metrics
            high_performance_weeks = sum(1 for week in weekly_data if week.get('revenue', 0) > 1000)
            medium_performance_weeks = sum(1 for week in weekly_data if 500 <= week.get('revenue', 0) <= 1000)
            low_performance_weeks = len(weekly_data) - high_performance_weeks - medium_performance_weeks
            
            # Add pie chart data starting at row 45
            chart_data_row = 45
            
            # Headers
            ws[f'A{chart_data_row}'] = "Performance Level"
            ws[f'B{chart_data_row}'] = "Number of Weeks"
            
            # Format headers
            for col in ['A', 'B']:
                cell = ws[f'{col}{chart_data_row}']
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.border
            
            # Add data
            pie_data = [
                ("High Performance (>1000 ETB)", high_performance_weeks),
                ("Medium Performance (500-1000 ETB)", medium_performance_weeks),
                ("Low Performance (<500 ETB)", low_performance_weeks)
            ]
            
            for i, (category, count) in enumerate(pie_data, 1):
                data_row = chart_data_row + i
                ws[f'A{data_row}'] = category
                ws[f'B{data_row}'] = count
                
                # Format cells
                for col in ['A', 'B']:
                    cell = ws[f'{col}{data_row}']
                    cell.font = self.data_font
                    cell.border = self.border
                    if col == 'B':
                        cell.number_format = '#,##0'
                        cell.alignment = self.center_alignment
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Create pie chart
            chart = PieChart()
            chart.title = "Weekly Performance Distribution"
            chart.width = 12
            chart.height = 10
            
            # Add data to chart
            data = Reference(ws, min_col=2, min_row=chart_data_row, max_row=chart_data_row + len(pie_data))
            labels = Reference(ws, min_col=1, min_row=chart_data_row + 1, max_row=chart_data_row + len(pie_data))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            
            # Position chart
            ws.add_chart(chart, "D45")
            
        except Exception as e:
            logger.error(f"Error creating performance pie chart: {e}", exc_info=True)
    
    def _create_summary_sheet(self, ws, weekly_data):
        """Create summary analysis sheet."""
        try:
            ws.title = "Summary"
            
            # Add summary title
            ws['A1'] = "Weekly Analytics Summary"
            ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='366092')
            ws.merge_cells('A1:D1')
            ws['A1'].alignment = self.center_alignment
            
            if not weekly_data:
                ws['A3'] = "No data available for summary analysis."
                return
            
            # Calculate summary statistics
            total_revenue = sum(week.get('revenue', 0) for week in weekly_data)
            total_new_users = sum(week.get('new_users', 0) for week in weekly_data)
            total_active_users = sum(week.get('active_users', 0) for week in weekly_data)
            total_messages = sum(week.get('total_messages', 0) for week in weekly_data)
            total_orders = sum(week.get('orders_requests', 0) for week in weekly_data)
            
            avg_weekly_revenue = total_revenue / len(weekly_data)
            avg_weekly_new_users = total_new_users / len(weekly_data)
            avg_weekly_messages = total_messages / len(weekly_data)
            
            # Best and worst weeks
            best_revenue_week = max(weekly_data, key=lambda x: x.get('revenue', 0))
            worst_revenue_week = min(weekly_data, key=lambda x: x.get('revenue', 0))
            most_active_week = max(weekly_data, key=lambda x: x.get('total_messages', 0))
            
            # Summary data
            summary_data = [
                ("Reporting Period", f"{len(weekly_data)} weeks"),
                ("", ""),
                ("REVENUE ANALYSIS", ""),
                ("Total Revenue", f"{total_revenue:.2f} ETB"),
                ("Average Weekly Revenue", f"{avg_weekly_revenue:.2f} ETB"),
                ("Best Week Revenue", f"{best_revenue_week.get('revenue', 0):.2f} ETB ({best_revenue_week.get('week_number', 'N/A')})"),
                ("Worst Week Revenue", f"{worst_revenue_week.get('revenue', 0):.2f} ETB ({worst_revenue_week.get('week_number', 'N/A')})"),
                ("", ""),
                ("USER ENGAGEMENT", ""),
                ("Total New Users", f"{total_new_users:,}"),
                ("Average Weekly New Users", f"{avg_weekly_new_users:.1f}"),
                ("Total Messages Sent", f"{total_messages:,}"),
                ("Average Weekly Messages", f"{avg_weekly_messages:.1f}"),
                ("Most Active Week", f"{most_active_week.get('total_messages', 0):,} messages ({most_active_week.get('week_number', 'N/A')})"),
                ("", ""),
                ("ORDER ACTIVITY", ""),
                ("Total Orders/Requests", f"{total_orders:,}"),
                ("Average Weekly Orders", f"{total_orders / len(weekly_data):.1f}"),
                ("", ""),
                ("GROWTH TRENDS", ""),
            ]
            
            # Calculate growth trends
            if len(weekly_data) >= 2:
                recent_avg = sum(week.get('revenue', 0) for week in weekly_data[:2]) / 2
                older_avg = sum(week.get('revenue', 0) for week in weekly_data[-2:]) / 2
                revenue_trend = ((recent_avg - older_avg) / older_avg * 100) if older_avg > 0 else 0
                
                recent_users_avg = sum(week.get('new_users', 0) for week in weekly_data[:2]) / 2
                older_users_avg = sum(week.get('new_users', 0) for week in weekly_data[-2:]) / 2
                users_trend = ((recent_users_avg - older_users_avg) / older_users_avg * 100) if older_users_avg > 0 else 0
                
                summary_data.extend([
                    ("Revenue Trend (Recent vs Older)", f"{revenue_trend:+.1f}%"),
                    ("User Acquisition Trend", f"{users_trend:+.1f}%")
                ])
            
            # Add summary data to worksheet
            for i, (metric, value) in enumerate(summary_data, 3):
                row = i
                
                if metric == "":
                    continue
                elif metric in ["REVENUE ANALYSIS", "USER ENGAGEMENT", "ORDER ACTIVITY", "GROWTH TRENDS"]:
                    # Section headers
                    ws[f'A{row}'] = metric
                    ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
                    ws[f'A{row}'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    ws.merge_cells(f'A{row}:D{row}')
                    ws[f'A{row}'].alignment = self.center_alignment
                    ws[f'A{row}'].border = self.border
                else:
                    # Regular data
                    ws[f'A{row}'] = metric
                    ws[f'C{row}'] = value
                    
                    # Format cells
                    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
                    ws[f'C{row}'].font = self.data_font
                    ws[f'A{row}'].border = self.border
                    ws[f'C{row}'].border = self.border
                    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                    ws[f'C{row}'].alignment = Alignment(horizontal='right', vertical='center')
            
            # Auto-adjust columns
            for col in ['A', 'B', 'C', 'D']:
                ws.column_dimensions[col].width = 25
            
        except Exception as e:
            logger.error(f"Error creating summary sheet: {e}", exc_info=True)
    
    def generate_weekly_comparison_report(self, db_manager, filename_prefix="Blue_Pharma_Weekly_Comparison"):
        """Generate weekly comparison Excel report with professional formatting and conditional formatting."""
        try:
            # Get comparison metrics
            comparison_metrics = db_manager.get_weekly_comparison_metrics()
            
            if not comparison_metrics:
                logger.warning("No comparison metrics data available")
                return None
            
            # Create temporary Excel file
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
                temp_file_path = temp_file.name
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Weekly Comparison"
            
            # Generate comparison sheet
            self._create_weekly_comparison_sheet(ws, comparison_metrics)
            
            # Save the workbook
            wb.save(temp_file_path)
            wb.close()
            
            # Create final filename with timestamp
            current_date = datetime.now().strftime('%Y-%m-%d_%H-%M')
            final_filename = f"{filename_prefix}_{current_date}.xlsx"
            
            return temp_file_path, final_filename
            
        except Exception as e:
            logger.error(f"Error generating weekly comparison report: {e}", exc_info=True)
            return None, None
    
    def _create_weekly_comparison_sheet(self, ws, comparison_metrics):
        """Create weekly comparison sheet with conditional formatting."""
        try:
            # Set sheet title
            ws['A1'] = "Blue Pharma Trading PLC - Weekly Comparison Report"
            ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='366092')
            ws.merge_cells('A1:H1')
            ws['A1'].alignment = self.center_alignment
            
            # Add generation date
            ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(name='Calibri', size=11, italic=True)
            ws.merge_cells('A2:H2')
            ws['A2'].alignment = self.center_alignment
            
            # Add period info
            current_week = datetime.now().strftime('%Y-%W')
            prev_week = (datetime.now() - timedelta(weeks=1)).strftime('%Y-%W')
            ws['A3'] = f"Comparison Period: Week {prev_week} vs Week {current_week}"
            ws['A3'].font = Font(name='Calibri', size=12, bold=True)
            ws.merge_cells('A3:H3')
            ws['A3'].alignment = self.center_alignment
            
            # Add some spacing
            ws.row_dimensions[4].height = 15
            
            # Create comparison table
            headers = [
                "Metric", "Previous Week", "Current Week", "Difference", "% Change", "Status", "Performance"
            ]
            
            # Add headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col_idx, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.border
            
            # Prepare comparison data
            comparison_data = []
            for metric_name, metric_data in comparison_metrics.items():
                current = metric_data.get('current', 0)
                previous = metric_data.get('previous', 0)
                
                # Calculate difference and percentage change
                difference = current - previous
                pct_change = ((current - previous) / previous * 100) if previous > 0 else 0
                
                # Determine status and performance
                if difference > 0:
                    status = "📈 Growth"
                    performance = "Positive"
                elif difference < 0:
                    status = "📉 Decline"
                    performance = "Negative"
                else:
                    status = "📊 Stable"
                    performance = "Neutral"
                
                # Format metric name
                formatted_metric = metric_name.replace('_', ' ').title()
                if 'Revenue' in formatted_metric:
                    formatted_metric = "Revenue (ETB)"
                elif 'Orders' in formatted_metric:
                    formatted_metric = "Orders/Requests"
                elif 'Messages' in formatted_metric:
                    formatted_metric = "Total Messages"
                elif 'Users' in formatted_metric and 'Active' in formatted_metric:
                    formatted_metric = "Active Users"
                elif 'Users' in formatted_metric and 'New' in formatted_metric:
                    formatted_metric = "New Users"
                
                comparison_data.append({
                    'metric': formatted_metric,
                    'previous': previous,
                    'current': current,
                    'difference': difference,
                    'pct_change': pct_change,
                    'status': status,
                    'performance': performance
                })
            
            # Add data rows
            for row_idx, data in enumerate(comparison_data, 6):
                ws.cell(row=row_idx, column=1, value=data['metric']).font = Font(name='Calibri', size=11, bold=True)
                ws.cell(row=row_idx, column=2, value=data['previous'])
                ws.cell(row=row_idx, column=3, value=data['current'])
                ws.cell(row=row_idx, column=4, value=data['difference'])
                ws.cell(row=row_idx, column=5, value=data['pct_change'])
                ws.cell(row=row_idx, column=6, value=data['status'])
                ws.cell(row=row_idx, column=7, value=data['performance'])
                
                # Format each cell
                for col_idx in range(1, 8):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = self.data_font
                    cell.border = self.border
                    
                    # Special formatting based on column
                    if col_idx in [2, 3]:  # Previous and Current values
                        if 'Revenue' in data['metric']:
                            cell.number_format = '#,##0.00'
                        else:
                            cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    elif col_idx == 4:  # Difference
                        if 'Revenue' in data['metric']:
                            cell.number_format = '+#,##0.00;-#,##0.00;0.00'
                        else:
                            cell.number_format = '+#,##0;-#,##0;0'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        
                        # Conditional formatting for difference
                        if data['difference'] > 0:
                            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                            cell.font = Font(name='Calibri', size=11, color='006100')
                        elif data['difference'] < 0:
                            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                            cell.font = Font(name='Calibri', size=11, color='9C0006')
                    elif col_idx == 5:  # Percentage change
                        cell.number_format = '+0.0%;-0.0%;0.0%'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        
                        # Conditional formatting for percentage change
                        if data['pct_change'] > 0:
                            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                            cell.font = Font(name='Calibri', size=11, color='006100', bold=True)
                        elif data['pct_change'] < 0:
                            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                            cell.font = Font(name='Calibri', size=11, color='9C0006', bold=True)
                    elif col_idx == 7:  # Performance
                        # Conditional formatting for performance
                        if data['performance'] == 'Positive':
                            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                            cell.font = Font(name='Calibri', size=11, color='006100', bold=True)
                        elif data['performance'] == 'Negative':
                            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                            cell.font = Font(name='Calibri', size=11, color='9C0006', bold=True)
                        else:
                            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                            cell.font = Font(name='Calibri', size=11, color='9C6500', bold=True)
                    else:
                        cell.alignment = self.center_alignment
            
            # Add insights section
            insights_row = len(comparison_data) + 8
            
            ws[f'A{insights_row}'] = "Weekly Performance Insights"
            ws[f'A{insights_row}'].font = Font(name='Calibri', size=14, bold=True, color='366092')
            ws.merge_cells(f'A{insights_row}:G{insights_row}')
            ws[f'A{insights_row}'].alignment = self.center_alignment
            
            # Generate insights
            insights = []
            
            # Calculate total revenue for insights
            total_revenue = sum(comparison_metrics.get(metric, {}).get('current', 0) + comparison_metrics.get(metric, {}).get('previous', 0) for metric in comparison_metrics if 'revenue' in metric.lower())
            
            # Revenue insights
            if total_revenue > 0:
                revenue_current = comparison_metrics.get('revenue', {}).get('current', 0)
                revenue_previous = comparison_metrics.get('revenue', {}).get('previous', 0)
                if revenue_current > revenue_previous:
                    insights.append("✅ Revenue is growing week-over-week")
                elif revenue_current < revenue_previous:
                    insights.append("⚠️ Revenue declined this week")
                else:
                    insights.append("📊 Revenue remained stable")
            
            # User activity insights
            new_users_current = comparison_metrics.get('new_users', {}).get('current', 0)
            active_users_current = comparison_metrics.get('active_users', {}).get('current', 0)
            
            if new_users_current > 10:
                insights.append("🎉 Strong user acquisition this week")
            elif new_users_current == 0:
                insights.append("⚠️ No new users acquired this week")
            
            if active_users_current > 50:
                insights.append("📈 High user engagement this week")
            elif active_users_current < 10:
                insights.append("📉 Low user engagement this week")
            
            # Add insights to worksheet
            for i, insight in enumerate(insights, insights_row + 2):
                ws[f'A{i}'] = insight
                ws[f'A{i}'].font = Font(name='Calibri', size=11)
                ws.merge_cells(f'A{i}:G{i}')
                ws[f'A{i}'].alignment = Alignment(horizontal='left', vertical='center')
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
            # Freeze header row
            ws.freeze_panes = 'A6'
            
        except Exception as e:
            logger.error(f"Error creating weekly comparison sheet: {e}", exc_info=True)
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content."""
        try:
            for column in ws.columns:
                max_length = 0
                column_letter = None
                
                # Get column letter from first non-merged cell
                for cell in column:
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                        break
                
                if not column_letter:
                    continue
                
                for cell in column:
                    try:
                        if hasattr(cell, 'value') and cell.value is not None:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set adjusted width with minimum and maximum limits
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.error(f"Error auto-adjusting columns: {e}", exc_info=True)
    
    def generate_weekly_analytics_report_from_sales(self, weekly_data, filename_prefix="Blue_Pharma_Weekly_Analytics"):
        """Generate weekly analytics report from sales data."""
        try:
            # Create temporary Excel file
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
                temp_file_path = temp_file.name
            
            # Create workbook and worksheets
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create multiple sheets
            ws_weekly = wb.create_sheet("Weekly Analytics")
            ws_dashboard = wb.create_sheet("Dashboard")
            ws_summary = wb.create_sheet("Summary")
            
            # Generate weekly analytics sheet
            self._create_weekly_analytics_sheet(ws_weekly, weekly_data)
            
            # Generate dashboard sheet with charts
            self._create_dashboard_sheet(ws_dashboard, weekly_data)
            
            # Generate summary sheet
            self._create_summary_sheet(ws_summary, weekly_data)
            
            # Save the workbook
            wb.save(temp_file_path)
            wb.close()
            
            # Create final filename with timestamp
            current_date = datetime.now().strftime('%Y-%m-%d_%H-%M')
            final_filename = f"{filename_prefix}_{current_date}.xlsx"
            
            return temp_file_path, final_filename
            
        except Exception as e:
            logger.error(f"Error generating weekly analytics report from sales: {e}", exc_info=True)
            return None, None
    
    def generate_weekly_comparison_report_from_data(self, comparison_metrics, filename_prefix="Blue_Pharma_Weekly_Comparison"):
        """Generate weekly comparison report from comparison metrics."""
        try:
            # Create temporary Excel file
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
                temp_file_path = temp_file.name
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Weekly Comparison"
            
            # Generate comparison sheet
            self._create_weekly_comparison_sheet(ws, comparison_metrics)
            
            # Save the workbook
            wb.save(temp_file_path)
            wb.close()
            
            # Create final filename with timestamp
            current_date = datetime.now().strftime('%Y-%m-%d_%H-%M')
            final_filename = f"{filename_prefix}_{current_date}.xlsx"
            
            return temp_file_path, final_filename
            
        except Exception as e:
            logger.error(f"Error generating weekly comparison report from data: {e}", exc_info=True)
            return None, None

# Global instance
excel_generator = ExcelAnalyticsGenerator()

def generate_enhanced_weekly_report(db_manager):
    """Generate enhanced weekly analytics report."""
    try:
        # Get weekly sales data first to check if we have any data
        weekly_sales = db_manager.get_weekly_sales_data(8)
        
        if not weekly_sales:
            return None, "No weekly sales data available for analytics report"
        
        # Convert sales data to analytics format
        weekly_data = []
        for week_data in weekly_sales:
            # Create analytics data from sales data
            analytics_data = {
                'week_number': week_data.get('week', ''),
                'week_start': '',  # We'll calculate this
                'week_end': '',    # We'll calculate this
                'new_users': 0,    # Default value since we don't have this data
                'active_users': 1, # Default value
                'total_messages': week_data.get('total_orders', 0) * 5,  # Estimate based on orders
                'orders_requests': week_data.get('total_orders', 0),
                'revenue': week_data.get('total_revenue', 0.0),
                'top_user': 'N/A',
                'notes': f"Revenue: {week_data.get('total_revenue', 0):.2f} ETB, Orders: {week_data.get('total_orders', 0)}"
            }
            weekly_data.append(analytics_data)
        
        if not weekly_data:
            return None, "No data available for weekly analytics report"
        
        result = excel_generator.generate_weekly_analytics_report_from_sales(weekly_data)
        if result is None:
            return None, "Error generating weekly analytics report"
        return result
        
    except Exception as e:
        logger.error(f"Error in generate_enhanced_weekly_report: {e}", exc_info=True)
        return None, f"Error generating report: {str(e)}"

def generate_enhanced_comparison_report(db_manager):
    """Generate enhanced weekly comparison report."""
    try:
        # Get weekly comparison data
        comparison_data = db_manager.get_weekly_comparison_data()
        
        if not comparison_data or len(comparison_data) < 2:
            return None, "Insufficient data for weekly comparison. Need at least 2 weeks of sales data."
        
        # Convert to metrics format
        current_week = comparison_data[0]
        previous_week = comparison_data[1] if len(comparison_data) > 1 else comparison_data[0]
        
        comparison_metrics = {
            'revenue': {
                'current': current_week.get('total_revenue', 0),
                'previous': previous_week.get('total_revenue', 0)
            },
            'new_users': {
                'current': 0,  # Default since we don't track this
                'previous': 0
            },
            'active_users': {
                'current': current_week.get('unique_customers', 0),
                'previous': previous_week.get('unique_customers', 0)
            },
            'total_messages': {
                'current': current_week.get('total_orders', 0) * 5,  # Estimate
                'previous': previous_week.get('total_orders', 0) * 5
            },
            'orders_requests': {
                'current': current_week.get('total_orders', 0),
                'previous': previous_week.get('total_orders', 0)
            }
        }
        
        result = excel_generator.generate_weekly_comparison_report_from_data(comparison_metrics)
        if result is None:
            return None, "Error generating weekly comparison report"
        return result
        
    except Exception as e:
        logger.error(f"Error in generate_enhanced_comparison_report: {e}", exc_info=True)
        return None, f"Error generating comparison report: {str(e)}"
